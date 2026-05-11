"""
excel_writer.py — Write daily and consolidated Excel workbooks.

Produces .xlsx files matching the MS Access output format.
All output paths are configurable via config.json.

SAVE PATHS (from config.json):
  daily_output_path        — Individual daily Purchase Plan files
  consolidated_output_path — Combined/comparison workbooks

USAGE:
  from excel_writer import write_daily_excel, write_consolidated_excel
"""

import os
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference


# ============================================================
# STYLING CONSTANTS
# ============================================================

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Format: Sub group | Itemcode | Desc | UnitPrc | [16 periods × 6-9 metrics]
PERIOD_METRICS = ['pr', 'pl', 'po', 'prod', 'svr', 'cb', 'cbAmt', 'ratio']
# Period 1 has extra 'OB' and 'sugg' (suggestion) columns
PERIOD1_METRICS = ['OB', 'sugg', 'pl', 'po', 'prod', 'svr', 'cb', 'cbAmt', 'ratio']
PERIOD2_METRICS = ['sugg', 'pl', 'po', 'prod', 'svr', 'cb', 'cbAmt', 'ratio']
PERIODN_METRICS = ['pr', 'prod', 'svr', 'cb', 'cbAmt', 'ratio']  # periods 3-16


def _apply_header_style(ws, row: int, ncols: int):
    """Apply standard header formatting to a row."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _auto_column_width(ws, max_width: int = 50):
    """Auto-size columns based on content."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = 8
        for cell in col:
            if cell.value:
                best = max(best, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = best + 2


# ============================================================
# DAILY EXCEL (replaces MS Access output)
# ============================================================

def write_daily_excel(
    items: List[Dict],
    period_dates: List[str],
    output_dir: str,
    date_str: str
) -> str:
    """
    Generate the daily Purchase Plan Excel file.

    Produces output matching the MS Access structure:
      - Sheet "Data": Clean flat data (item-level, all periods)
      - Sheet "PivotView": Pivot-style layout matching Access Sheet4

    Args:
        items:         Aggregated item rows from text_reader
        period_dates:  Sorted period date strings (YYYYMMDD)
        output_dir:    SAVE PATH — where to write the .xlsx file
                       (from config.json → daily_output_path)
        date_str:      Date string YYYYMMDD for filename

    Returns:
        Full path to the generated file
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Data (clean flat format) ──
    ws_data = wb.active
    ws_data.title = "Data"

    # Build headers
    fixed_headers = [
        'ItemCode', 'Description', 'OldMaterial', 'MRPController',
        'MaterialType', 'MaterialGroup', 'ProfitCenter', 'UoM',
        'UnitPrice', 'SourceEntity'
    ]
    balance_headers = ['BL_PR', 'BL_PO', 'BL_PL', 'BL_NORM', 'BL_SVR']

    period_headers = []
    for pd in period_dates:
        for suffix in ['PR', 'PO', 'PL', 'NORM', 'SVR', 'ESTCB', 'AMOUNT', 'RATIO']:
            period_headers.append(f"{pd}_{suffix}")

    all_headers = fixed_headers + balance_headers + period_headers
    _apply_header_style(ws_data, 1, len(all_headers))
    for c, h in enumerate(all_headers, 1):
        ws_data.cell(row=1, column=c, value=h)

    for i, item in enumerate(items):
        r = i + 2
        # Fixed columns
        for c, h in enumerate(fixed_headers, 1):
            ws_data.cell(row=r, column=c, value=item.get(h, ''))

        # Balance columns
        col_offset = len(fixed_headers) + 1
        for c, h in enumerate(balance_headers, col_offset):
            ws_data.cell(row=r, column=c, value=item.get(h, 0.0))

        # Period columns
        col_offset = len(fixed_headers) + len(balance_headers) + 1
        for c, h in enumerate(period_headers, col_offset):
            ws_data.cell(row=r, column=c, value=item.get(h, 0.0))

    ws_data.freeze_panes = 'A2'
    ws_data.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(items)+1}"
    _auto_column_width(ws_data)

    # ── Sheet 2: PivotView (mimics Access Sheet4 pivot) ──
    ws_pivot = wb.create_sheet("PivotView")

    # Build period column names matching Access format
    # Access labels: 1st, 2nd, 3rd...16th
    ordinal_suffixes = {1: '1st', 2: '2nd', 3: '3rd'}
    for n in range(4, 21):
        ordinal_suffixes[n] = f'{n}th'

    # Match Access Sheet4 header row exactly: 'Sub group' (with space),
    # 'Itemcode' (lowercase c), and only 3 balance columns (no blnorm/blsvr).
    pivot_cols_base = ['Sub group', 'Itemcode', 'Desc', 'UnitPrc',
                       'blpr', 'blpo', 'blpl']
    pivot_cols_period = []

    for pn in range(1, min(len(period_dates) + 1, 17)):  # up to 16 periods
        pre = ordinal_suffixes.get(pn, f'{pn}th')
        if pn == 1:
            metrics = PERIOD1_METRICS
        elif pn == 2:
            metrics = PERIOD2_METRICS
        else:
            metrics = PERIODN_METRICS
        for m in metrics:
            pivot_cols_period.append(f"{pre}{m}")

    pivot_headers = pivot_cols_base + pivot_cols_period
    _apply_header_style(ws_pivot, 1, len(pivot_headers))
    for c, h in enumerate(pivot_headers, 1):
        ws_pivot.cell(row=1, column=c, value=h)

    # Map text periods to Excel pivot periods
    for i, item in enumerate(items):
        r = i + 2

        # Base columns. "Sub group" is the first whitespace-separated token
        # of Description (e.g. 'HBNX30 CONNECTOR CAP...' -> 'HBNX30'). Access
        # populates this on the first row only; we fill every row for
        # usable filtering/grouping.
        desc = item.get('Description', '')
        sub_group = desc.split(None, 1)[0] if desc.strip() else ''
        ws_pivot.cell(row=r, column=1, value=sub_group)
        ws_pivot.cell(row=r, column=2, value=item.get('ItemCode', ''))
        ws_pivot.cell(row=r, column=3, value=desc)
        ws_pivot.cell(row=r, column=4, value=item.get('UnitPrice', 0.0))

        # Balance (3 cols, matching Access: blpr, blpo, blpl)
        for ci, bl_key in enumerate(['BL_PR', 'BL_PO', 'BL_PL']):
            ws_pivot.cell(row=r, column=5 + ci, value=item.get(bl_key, 0.0))

        # Period data — first period starts at column 8 (4 identity + 3 balance + 1)
        pc = 8
        for pn, pd in enumerate(period_dates[:16]):  # max 16 periods
            text_map = {
                'pr': item.get(f'{pd}_PR', 0.0),
                'pl': item.get(f'{pd}_PL', 0.0),
                'po': item.get(f'{pd}_PO', 0.0),
                'prod': item.get(f'{pd}_NORM', 0.0),
                'svr': item.get(f'{pd}_SVR', 0.0),
                'cb': item.get(f'{pd}_ESTCB', 0.0),
                'cbAmt': item.get(f'{pd}_AMOUNT', 0.0),
                'ratio': item.get(f'{pd}_RATIO', 0.0),
            }

            if pn == 0:  # 1st period
                for m in PERIOD1_METRICS:
                    if m in ('OB', 'sugg'):
                        ws_pivot.cell(row=r, column=pc, value=0.0)
                    else:
                        ws_pivot.cell(row=r, column=pc, value=text_map.get(m, 0.0))
                    pc += 1
            elif pn == 1:  # 2nd period
                for m in PERIOD2_METRICS:
                    key = 'pr' if m == 'sugg' else m
                    ws_pivot.cell(row=r, column=pc, value=text_map.get(key, 0.0))
                    pc += 1
            else:  # 3rd-16th
                for m in PERIODN_METRICS:
                    ws_pivot.cell(row=r, column=pc, value=text_map.get(m, 0.0))
                    pc += 1

    ws_pivot.freeze_panes = 'E2'
    ws_pivot.auto_filter.ref = f"A1:{get_column_letter(len(pivot_headers))}{len(items)+1}"
    _auto_column_width(ws_pivot)

    # Save
    filename = f"PurchasePlan_{date_str}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ============================================================
# CONSOLIDATED EXCEL (multi-date comparison)
# ============================================================

def write_consolidated_excel(
    all_data: List[Dict],
    dates: List[str],
    period_keys: List[str],
    output_dir: str,
    month_label: str = "Consolidated"
) -> str:
    """
    Generate the consolidated comparison workbook.

    Creates 4 sheets:
      1. All Data — every item × every date (filterable)
      2. Summary by Date — daily totals + trend chart
      3. Day Changes — day-over-day deltas (green/red)
      4. Top Monthly Changes — biggest movers across the full period

    Args:
        all_data:      List of rows, each with 'Date', 'ItemCode', etc.
        dates:         Sorted unique date strings
        period_keys:   Which period metrics to compare (e.g., ['1stpo', '1stcb'])
        output_dir:    SAVE PATH — where to write the file
                       (from config.json → consolidated_output_path)
        month_label:   Label for filename (e.g., "Feb2026")

    Returns:
        Full path to generated file
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()

    # Build date-item index for fast lookup
    date_item_index = {}
    for i, row in enumerate(all_data):
        key = (row['DateSort'], row['ItemCode'])
        date_item_index[key] = i

    def get_metrics(date_str, item_code):
        idx = date_item_index.get((date_str, item_code))
        if idx is None:
            return None
        row = all_data[idx]
        return {k: float(row.get(k, 0) or 0) for k in period_keys}

    dates_disp = sorted(set(d['Date'] for d in all_data))

    # ── Sheet 1: All Data ──
    ws1 = wb.active
    ws1.title = "All Data"
    fixed = ['Date', 'SubGroup', 'ItemCode', 'Desc', 'UnitPrc']
    period_hdrs = [h for h in all_data[0].keys()
                   if h not in fixed + ['DateSort', 'SourceEntity']]
    all_h = fixed + sorted(period_hdrs)
    _apply_header_style(ws1, 1, len(all_h))
    for c, h in enumerate(all_h, 1):
        ws1.cell(row=1, column=c, value=h)
    for i, row in enumerate(all_data):
        r = i + 2
        for c, h in enumerate(all_h, 1):
            ws1.cell(row=r, column=c, value=row.get(h, ''))
    ws1.freeze_panes = 'F2'
    _auto_column_width(ws1)

    # ── Sheet 2: Summary by Date ──
    ws2 = wb.create_sheet("Summary by Date")
    sh = ['Date', 'Items'] + period_keys
    _apply_header_style(ws2, 1, len(sh))
    for c, h in enumerate(sh, 1):
        ws2.cell(row=1, column=c, value=h)

    date_totals = {}
    for ds in dates:
        t = {k: 0.0 for k in period_keys}
        t['n'] = 0
        for (d, item) in date_item_index:
            if d == ds:
                row = all_data[date_item_index[(d, item)]]
                for k in period_keys:
                    t[k] += float(row.get(k, 0) or 0)
                t['n'] += 1
        date_totals[ds] = t

    for i, ds in enumerate(dates):
        r = i + 2
        ws2.cell(row=r, column=1, value=f"{ds[4:6]}/{ds[6:8]}")
        ws2.cell(row=r, column=2, value=date_totals[ds]['n'])
        for c, k in enumerate(period_keys, 3):
            ws2.cell(row=r, column=c, value=round(date_totals[ds][k], 2))

    # Trend chart
    chart = LineChart()
    chart.title = "Daily Totals"
    chart.style = 10
    chart.height = 14
    chart.width = 22
    chart.add_data(Reference(ws2, min_col=3, min_row=1,
                   max_col=2+len(period_keys), max_row=len(dates)+1),
                   titles_from_data=True)
    chart.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=len(dates)+1))
    for s in chart.series:
        s.graphicalProperties.line.width = 25000
    ws2.add_chart(chart, "A20")
    ws2.freeze_panes = 'A2'
    _auto_column_width(ws2)

    # ── Sheet 3: Day Changes ──
    ws3 = wb.create_sheet("Day Changes")
    changes = []
    for i in range(1, len(dates)):
        pd_s, cd_s = dates[i-1], dates[i]
        pd_disp, cd_disp = dates_disp[i-1], dates_disp[i]
        curr_items = set(item for (d, item) in date_item_index if d == cd_s)
        for item in curr_items:
            pm = get_metrics(pd_s, item)
            cm = get_metrics(cd_s, item)
            if pm is None or cm is None:
                continue
            deltas = {k: cm[k] - pm[k] for k in period_keys}
            if any(abs(v) > 0.01 for v in deltas.values()):
                idx = date_item_index[(cd_s, item)]
                rd = all_data[idx]
                ch = {'From': pd_disp, 'To': cd_disp,
                      'ItemCode': item, 'Desc': rd['Desc'],
                      'SubGroup': rd['SubGroup']}
                for k in period_keys:
                    ch[f'Prev_{k}'] = pm[k]
                    ch[f'Curr_{k}'] = cm[k]
                    ch[f'Delta_{k}'] = deltas[k]
                changes.append(ch)

    ch_h = ['From', 'To', 'ItemCode', 'Desc', 'SubGroup']
    for k in period_keys:
        ch_h += [f'Prev_{k}', f'Curr_{k}', f'Delta_{k}']
    _apply_header_style(ws3, 1, len(ch_h))
    for c, h in enumerate(ch_h, 1):
        ws3.cell(row=1, column=c, value=h)

    for i, ch in enumerate(changes):
        r = i + 2
        for c, h in enumerate(ch_h, 1):
            v = ch.get(h, '')
            cell = ws3.cell(row=r, column=c,
                            value=round(v, 2) if isinstance(v, float) else v)
            if h.startswith('Delta_') and isinstance(v, float) and abs(v) > 0.01:
                cell.fill = POSITIVE_FILL if v > 0 else NEGATIVE_FILL
    ws3.freeze_panes = 'F2'
    _auto_column_width(ws3)

    # ── Sheet 4: Top Monthly Changes ──
    ws4 = wb.create_sheet("Top Monthly Changes")
    fd, ld = dates[0], dates[-1]
    ic = []
    for item in set(item for (d, item) in date_item_index):
        fm = get_metrics(fd, item)
        lm = get_metrics(ld, item)
        if fm is None or lm is None:
            continue
        da = lm.get('1stcbAmt', lm.get(list(period_keys)[0], 0)) - \
             fm.get('1stcbAmt', fm.get(list(period_keys)[0], 0))
        if abs(da) > 0.01:
            idx = date_item_index.get((ld, item)) or date_item_index.get((fd, item))
            rd = all_data[idx]
            ic.append({
                'ItemCode': item, 'Desc': rd['Desc'], 'SubGroup': rd['SubGroup'],
                'FirstAmt': fm.get('1stcbAmt', fm.get(list(period_keys)[0], 0)),
                'LastAmt': lm.get('1stcbAmt', lm.get(list(period_keys)[0], 0)),
                'DeltaAmt': da, 'AbsChange': abs(da),
            })
    ic.sort(key=lambda x: x['AbsChange'], reverse=True)

    th = ['ItemCode', 'Desc', 'SubGroup', 'FirstAmt', 'LastAmt', 'DeltaAmt', 'AbsChange']
    _apply_header_style(ws4, 1, len(th))
    for c, h in enumerate(th, 1):
        ws4.cell(row=1, column=c, value=h)
    for i, ch in enumerate(ic[:500]):
        r = i + 2
        for c, h in enumerate(th, 1):
            v = ch[h]
            cell = ws4.cell(row=r, column=c,
                            value=round(v, 2) if isinstance(v, float) else v)
            if h == 'DeltaAmt' and isinstance(v, float):
                cell.fill = POSITIVE_FILL if v > 0 else NEGATIVE_FILL if v < 0 else None
    ws4.freeze_panes = 'A2'
    _auto_column_width(ws4)

    # Save
    filename = f"PP29_{month_label}_Consolidated.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
